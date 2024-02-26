import openpyxl as xl
from openpyxl.styles import PatternFill
import json
import requests

weather_url = 'https://danepubliczne.imgw.pl/api/data/synop'
weather_excel = xl.open('weather_station.xlsx')     # workbook object
sheet_1 = weather_excel['Arkusz1']  # Sheet1


def write_1st_row(dictionary):
    i = 1
    for key, value in dictionary.items():
        cell = sheet_1.cell(1, i)
        cell.value = key
        i += 1
    weather_excel.save('weather_station.xlsx')


def read_json_to_excel():
    response = requests.get(weather_url)
    if response.status_code == 200:
        # Parse JSON
        data = json.loads(response.text)
        write_1st_row(data[0])
        row, col = 2, 1
        for station in data:
            for key, value in station.items():
                cell2 = sheet_1.cell(row, col)
                cell2.value = value
                if value is None:
                    redFill = PatternFill(
                        start_color='ff0000',
                        end_color='ff0000',
                        fill_type='solid'
                    )
                    cell2.fill = redFill
                    cell2.value = 'Brak Danych'
                col += 1
            col = 1
            row += 1
        weather_excel.save('weather_station.xlsx')
    else:
        print("Failed to retrieve data: ", response.status_code)


if __name__ == "__main__":
    read_json_to_excel()
